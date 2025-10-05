import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import io
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------
#  HIDE STREAMLIT UI ELEMENTS
# ------------------------------------------------------------------
st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    </style>
""", unsafe_allow_html=True)

# ------------------------------------------------------------------
#  LOGIN PAGE
# ------------------------------------------------------------------
def login_page():
    st.markdown(""" 
        <style>
        .login-container {
            max-width: 280px;
            margin: 60px auto;
            padding: 15px 20px;
            background: #f0f2f6;
            border-radius: 6px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.1);
        }
        .login-header {
            font-size: 20px;
            font-weight: 600;
            color: #333;
            margin-bottom: 15px;
            text-align: center;
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        }
        </style>
        <div class="login-container">
            <div class="login-header">Please Log In</div>
        </div>
    """, unsafe_allow_html=True)

    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Login")

    if submitted:
        if username in st.secrets.get("auth", {}) and password == st.secrets["auth"][username]:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("Invalid username or password.")

# ------------------------------------------------------------------
#  AUTH GATE
# ------------------------------------------------------------------
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

if not st.session_state["authenticated"]:
    login_page()
    st.stop()

# ------------------------------------------------------------------
#  SIDEBAR LOGOUT
# ------------------------------------------------------------------
with st.sidebar:
    if st.button("Logout"):
        st.session_state["authenticated"] = False
        st.rerun()

# ------------------------------------------------------------------
#  REPORT ENGINE
# ------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def build_report(file_bytes):
    TODAY     = pd.Timestamp.today().normalize()
    YESTERDAY = TODAY - timedelta(days=1)

    df = pd.read_excel(io.BytesIO(file_bytes))
    # normalise headers
    df.columns = (df.columns
                  .str.strip()
                  .str.replace('\u00A0',' ')
                  .str.lower()
                  .str.replace(' ','_'))

    df['date'] = pd.to_datetime(df['date'], dayfirst=True)
    df['interest'] = pd.to_numeric(df['interest'], errors='coerce')

    calc_df = df[df['nature'].str.upper().eq('CALC')].copy().sort_values('date')
    post_df = df[df['nature'].str.upper().eq('POST')].copy().sort_values('date')

    if calc_df.empty or post_df.empty:
        return pd.DataFrame([{
            'Due Date': 'No CALC or POST rows',
            'Interest Due': 0,
            'Paid Dates': '',
            'Amount Paid': 0,
            'Balance Due': 0,
            'Overdue_Days': '',
            'Status': ''
        }])

    payments  = (-post_df['interest']).tolist()
    pay_dates = post_df['date'].dt.date.tolist()

    # 10th schedule
    start, end = calc_df['date'].min(), calc_df['date'].max()
    first_10th = datetime(start.year,
                          start.month if start.day >= 10 else start.month - 1,
                          10)
    all_10ths = pd.date_range(first_10th, end, freq='MS').shift(9, freq='D')

    monthly_rows = []
    for due_10 in all_10ths:
        prev_11 = (due_10.replace(day=11) - timedelta(days=30)).replace(day=11)
        mask = (calc_df['date'] >= prev_11) & (calc_df['date'] <= due_10)
        monthly_rows.append({
            'due_date': due_10.date(),
            'interest_due': round(calc_df.loc[mask, 'interest'].sum(), 2)
        })

    # apply payments
    report_lines = []
    rem_pays, rem_dates = payments.copy(), pay_dates.copy()

    for row in monthly_rows:
        due_dt, due_amt = row['due_date'], row['interest_due']
        paid, used = 0.0, []
        while due_amt > 1e-2 and rem_pays:
            p, d = rem_pays.pop(0), rem_dates.pop(0)
            used.append(str(d))
            if p >= due_amt:
                rem_pays.insert(0, p - due_amt)
                rem_dates.insert(0, d)
                paid += due_amt
                due_amt = 0
                break
            else:
                paid += p
                due_amt -= p

        balance = round(row['interest_due'] - paid, 2)
        status  = '✅ Fully Paid' if balance < 1e-2 else '❌ Outstanding'
        overdue_days = max(0, (YESTERDAY - pd.Timestamp(due_dt)).days) if balance >= 1e-2 else ''

        report_lines.append({
            'Due Date': due_dt,
            'Interest Due': round(row['interest_due'], 2),
            'Paid Dates': ' || '.join(used) if used else '—',
            'Amount Paid': round(paid, 2),
            'Balance Due': balance,
            'Overdue_Days': overdue_days,
            'Status': status
        })

    # accrued after last 10th
    last_10 = all_10ths[-1]
    accrued_mask = calc_df['date'] > last_10
    if accrued_mask.any():
        ai = calc_df.loc[accrued_mask, 'interest'].sum()
        report_lines.append({
            'Due Date': f"{(last_10 + pd.Timedelta(days=1)).strftime('%d-%b-%Y')} → {calc_df['date'].max().strftime('%d-%b-%Y')} (accrued)",
            'Interest Due': round(ai, 2),
            'Paid Dates': '',
            'Amount Paid': 0,
            'Balance Due': round(ai, 2),
            'Overdue_Days': '',
            'Status': 'Interest Accrued (not yet due)'
        })

    # total row
    report_lines.append({
        'Due Date': 'Total',
        'Interest Due': round(sum(r['Interest Due'] for r in report_lines[:-1]), 2),
        'Paid Dates': '',
        'Amount Paid': round(sum(r['Amount Paid'] for r in report_lines[:-1]), 2),
        'Balance Due': round(sum(r['Balance Due'] for r in report_lines[:-1]), 2),
        'Overdue_Days': '',
        'Status': ''
    })

    return pd.DataFrame(report_lines)

# ------------------------------------------------------------------
#  EXCEL STYLER
# ------------------------------------------------------------------
def style_excel(df: pd.DataFrame) -> io.BytesIO:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Report')
        ws = writer.sheets['Report']

        # formats
        bold = Font(bold=True)
        currency = u'#,##0.00_);[Red](#,##0.00)'
        left_align = Alignment(horizontal='left')

        # column widths
        for idx, col in enumerate(df.columns, 1):
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len, 50)

        # styling
        for row in range(2, len(df) + 2):
            ws.cell(row, 1).alignment = left_align   # Due Date
            for col_idx in [2, 4, 5]:                # Interest, Paid, Balance
                cell = ws.cell(row, col_idx)
                cell.number_format = currency
                if df.iloc[row-2, 0] == 'Total':
                    cell.font = bold
        # bold total row
        for col in range(1, len(df.columns)+1):
            ws.cell(len(df)+1, col).font = bold

    buffer.seek(0)
    return buffer

# ------------------------------------------------------------------
#  STREAMLIT UI
# ------------------------------------------------------------------
st.set_page_config(page_title='Payment Due Report', layout='centered')

st.markdown("""
<style>
    .block-container { max-width: 720px; margin: auto; }
    .main-header { background:#0e1117; padding:1.2rem 0; border-radius:12px; margin-bottom:1.5rem; }
    .main-header h2 { color:#fafafa; text-align:center; margin:0; font-family:'Source Sans Pro',sans-serif; }
    .css-1kyxreq.e1fqkh3o2 { background:#ffffff; border-radius:12px; padding:2rem 2.5rem; box-shadow:0 2px 6px rgba(0,0,0,.08); }
    .stDownloadButton > button {
        background:#0066cc; color:white; border:none; border-radius:8px; padding:0.5rem 1.5rem; font-weight:600;
    }
    .stDownloadButton > button:hover { background:#0052a3; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-header"><h2>Payment Due Report Generator</h2></div>', unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    'Upload Excel files (Sheet1)',
    type=['xlsx'],
    accept_multiple_files=True
)

if uploaded_files:
    if len(uploaded_files) > 10:
        st.error('You can upload a maximum of 10 files at once.')
        st.stop()

    for upl in uploaded_files:
        try:
            df_out = build_report(upl.getvalue())
            excel_bytes = style_excel(df_out)
            st.download_button(
                label=f'📥 {Path(upl.name).stem}_Payment_Due_Report.xlsx',
                data=excel_bytes,
                file_name=f'{Path(upl.name).stem}_Payment_Due_Report.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            st.error(f'Processing {upl.name} failed: {e}')
